from collections import namedtuple
from datetime import datetime
from itertools import izip, takewhile
from django.conf import settings
from django.utils import timezone
from openpyxl import load_workbook

from django.db import transaction
import pysftp
import pytz

from django.contrib.auth.models import User

from echo.apps.projects.models import Language, Project, VoiceSlot, VUID, UpdateStatus
from echo.apps.activity.models import Action
from echo.apps.projects.tasks import update_file_statuses


PAGE_NAME = "page name"
PROMPT_NAME = "prompt name"
PROMPT_TEXT = "prompt text"
LANGUAGE = "language"
STATE_NAME = "state name"
DATE_CHANGED = "date changed"

VUID_HEADER_NAME_SET = {
    PROMPT_NAME,
    PROMPT_TEXT,
    DATE_CHANGED
}


class FileStatus(object):
    def __init__(self, pathname, msum, mtime):
        self.pathname = pathname
        self.msum = msum
        self.mtime = float(mtime)


def allnamesequal(name):
    return all(n == name[0] for n in name[1:])


def commonprefix(paths, sep='/'):
    bydirectorylevels = zip(*[p.split(sep) for p in paths])
    return sep.join(x[0] for x in takewhile(allnamesequal, bydirectorylevels))

@transaction.atomic
def update_checksum(project, user):
    project = Project.objects.get(pk=int(project))
    # user = User.objects.get(pk=int(user))
    try:
        with pysftp.Connection(project.bravo_server.address, username=project.bravo_server.account,
                               private_key=settings.PRIVATE_KEY) as sftp:
            result = sftp.execute('find {0}/ -name "*.wav"'.format(project.root_path) +
                                  ' -exec md5sum {} \; -exec stat -c"%Y" {} \;')
    except IOError:
        # something in the execute didn't stir the kool-aid
        return {"valid": False, "message": "Error running command on server"}

    FileStatus = namedtuple('FileStatus', 'md5 path modified')
    file_statuses = []
    try:
        for i in range(0, len(result), 2):
            md5 = result[i].split()[0]
            filename = ' '.join(result[i].split()[1:])
            modified = result[i + 1].strip()
            file_statuses.append(FileStatus(md5, filename, modified))
    except IndexError:
        print "Update IndexError Result:"
        print repr(result)

    slots = project.voiceslots()

    for fs in file_statuses:
        if slots.filter(name=fs.path.split('/')[-1][:-4]).exists():
            slot_candidates = slots.filter(name=fs.path.split('/')[-1][:-4])
            for slot in slot_candidates:
                if fs.path == slot.filepath():
                    if slot.status in (VoiceSlot.PASS, VoiceSlot.FAIL):
                        #if fs.md5 != slot.bravo_checksum:
                            Action.log(user, Action.UPDATE_CHECKSUM, slot.bravo_checksum, slot)
                            slot.bravo_checksum = fs.md5
                            slot.bravo_time = datetime.utcfromtimestamp(float(fs.modified)).replace(tzinfo=pytz.utc)
                            slot.save()
                            Action.log(user, Action.UPDATE_CHECKSUM, fs.md5, slot)


@transaction.atomic
def fetch_slots_from_server(project, sftp, user):
    """Contains logic to update file statuses"""
    # get shared path of all distinct paths from voiceslot models in project
    # path = commonprefix(project.voiceslots().values_list('path', flat=True).distinct())
    path = project.root_path
    try:
        result = sftp.execute('find {0}/ -name "*.wav"'.format(path) + ' -exec md5sum {} \; -exec stat -c"%Y" {} \;')
    except IOError:
        # something in the execute didn't stir the kool-aid
        return {"valid": False, "message": "Error running command on server"}
    if len(result) == 0:
        # means path exists, but no files in path, mark all files as missing
        slots = project.voiceslots()
        for slot in slots:
            slot.status = VoiceSlot.MISSING
            slot.save()
            Action.log(user, Action.AUTO_MISSING_SLOT, 'Slot found missing during status check', slot)


        return {"valid": False, "message": "All files missing on server, given path \"{0}\"".format(path)}
    elif len(result) == 1 and result[0].startswith('find:'):
        # means error was returned, path does not exist
        return {"valid": False, "message": "Path \"{0}\" does not exist on server".format(path)}
    else:
        if len(result) % 2 == 1:
            # dataset should be 2 lines per record, if odd then something is not right
            return {"valid": False, "message": "Server providing invalid dataset"}
        else:
            # parse result into dictionary using izip
            l = izip(*([iter(result)]*2))
            map = {}
            for i in l:
                msum, pathname = i[0].strip().split('  ')
                mtime = i[1].strip()
                map[pathname] = FileStatus(pathname, msum, mtime)
            # get voiceslots for project and iterate over them
            slots = project.voiceslots()
            for slot in slots:
                # check for slot.filepath() in map.keys()
                if slot.filepath() not in map:
                    # if slot not in map, slot is missing
                    slot.status = VoiceSlot.MISSING
                    slot.save()
                    Action.log(user, Action.AUTO_MISSING_SLOT, 'Slot found missing during status check', slot)

                else:
                    # else slot in map, run additional tests
                    fs = map.get(slot.filepath())
                    bravo_time = datetime.fromtimestamp(fs.mtime)
                    if slot.bravo_time:
                        bravo_time = slot.bravo_time.replace(tzinfo=None)
                    if slot.status == VoiceSlot.MISSING:
                        slot.status = VoiceSlot.READY
                        slot.bravo_checksum = fs.msum
                        slot.bravo_time = timezone.make_aware(datetime.fromtimestamp(fs.mtime), timezone.get_current_timezone())
                        slot.save()
                        Action.log(user, Action.AUTO_NEW_SLOT, 'Slot discovered during status check', slot)

                    elif slot.bravo_time is None or bravo_time < datetime.fromtimestamp(fs.mtime) and slot.bravo_checksum != fs.msum:
                        slot.status = VoiceSlot.READY
                        slot.bravo_checksum = fs.msum
                        slot.bravo_time = timezone.make_aware(datetime.fromtimestamp(fs.mtime), timezone.get_current_timezone())
                        slot.save()
                        Action.log(user, Action.AUTO_NEW_SLOT, 'Slot discovered during status check', slot)

            return {"valid": True, "message": "Files from Bravo Server have been fetched"}


def make_filename(path, name):
    if path.endswith('/') and name.startswith('/'):
        return "{0}{1}".format(path[:-1], name)
    elif path.endswith('/') or name.startswith('/'):
        return "{0}{1}".format(path, name)
    return "{0}/{1}".format(path, name)


@transaction.atomic
def parse_vuid(vuid):
    wb = load_workbook(vuid.file.path)
    ws = wb.active

    headers = [str(i.value).lower() for i in ws.rows[0]]
    try:
        prompt_name_i = headers.index(PROMPT_NAME)
        prompt_text_i = headers.index(PROMPT_TEXT)
        date_changed_i = headers.index(DATE_CHANGED)
    except ValueError:
        return {"valid": False, "message": "Parser error, invalid headers"}

    no_language = False
    try:
        language_i = headers.index(LANGUAGE)
    except ValueError:
        no_language = True

    v = unicode(ws['A2'].value).strip()
    # if endswith '/', remove it
    if v[-1] == '/':
        v = v[:-1]
    i = v.find('/')
    path = v[i:].strip()
    slots = []

    for w in ws.rows[2:]:
        try:
            if no_language:
                language = Language.objects.get(project=vuid.project, name=unicode('english'))
            elif w[language_i].value is not None:
                language = Language.objects.get(project=vuid.project, name=unicode(w[language_i].value).strip().lower())
            else:
                language = Language.objects.get(project=vuid.project, name=unicode('english'))
        except Language.DoesNotExist:
            if no_language:
                language = Language(project=vuid.project, name=unicode('english'))
            else:
                language = Language(project=vuid.project, name=unicode(w[language_i].value).strip().lower())
            language.save()
        except Language.MultipleObjectsReturned:
            return {"valid": False, "message": "Parser error, multiple languages returned"}

        name = unicode(w[prompt_name_i].value).strip()
        verbiage = unicode(w[prompt_text_i].value).strip()
        vuid_time = w[date_changed_i].value.date() if w[date_changed_i].value is datetime else None

        try:
            vs = VoiceSlot.objects.get(name=name, path=path, language=language)
            if vuid_time > vs.vuid_time:
                vs.vuid_time = vuid_time.date()
                vs.verbiage = verbiage
                vs.vuid = vuid
            elif vuid_time == vs.vuid_time:
                if vs.verbiage != verbiage:
                    vs.verbiage = verbiage
                    vs.vuid = vuid
            slots.append(vs)
            vs.save()
        except VoiceSlot.DoesNotExist:
            vs = VoiceSlot(name=name, path=path, verbiage=verbiage, language=language, vuid_time=vuid_time, vuid=vuid)
            slots.append(vs)
            vs.save()
        except VoiceSlot.MultipleObjectsReturned:
            return {"valid": False, "message": "Parser error, multiple voice slots returned"}
    return {"valid": True, "message": "Parsed file successfully"}


def upload_vuid(uploaded_file, user, project):
    # check if project root path is set
    if not project.root_path:
        return {"valid": False, "message": "Please set root path, unable to upload"}

    vuid = VUID(filename=uploaded_file.name, file=uploaded_file, project=project, upload_by=user)
    vuid.save()

    # check if any cell of 'vuid file header' is empty
    if not verify_vuid_headers_empty(vuid):
        vuid.delete()
        return {"valid": False, "message": "Invalid file structure, unable to upload"}

    # check conflict between root path and vuid path
    if not verify_root_path(vuid):
        vuid.delete()
        return {"valid": False, "message": "Invalid vuid path, unable to upload"}

    result = verify_vuid(vuid)
    if not result['valid']:
        vuid.delete()
        return result
    result = parse_vuid(vuid)
    if not result['valid']:
        vuid.delete()
        return result
    Action.log(user, Action.UPLOAD_VUID, 'Prompt list {0} uploaded'.format(uploaded_file.name), project)

    if project.status == Project.TESTING:
        # set project status to "Initial"
        project.status = Project.INITIAL
        project.save()
        Action.log(user, Action.PROJECT_RECALLED, 'Project is recalled', project)

    status = UpdateStatus.objects.get_or_create(project=project)[0]
    query_item = update_file_statuses.delay(project_id=project.pk, user_id=user.id)
    status.query_id = query_item
    status.running = True
    status.save()
    return {"valid": True, "message": "File uploaded and parsed successfully"}


def verify_vuid(vuid):
    wb = load_workbook(vuid.file.path)
    ws = wb.active
    valid = False
    message = "Invalid file structure, unable to upload"
    if len(ws.rows) > 2:
        if not verify_vuid_headers(vuid):
            message = "Invalid file headers, unable to upload"
        elif not verify_root_path(vuid):
            message = "Invalid vuid path, unable to upload"
        else:
            valid = True
            message = "Uploaded file successfully"
    elif len(ws.rows) == 2:
        if verify_vuid_headers(vuid):
            message = "No records in file, unable to upload"
    return {"valid": valid, "message": message}


def verify_vuid_headers(vuid):
    wb = load_workbook(vuid.file.path)
    ws = wb.active
    if len(ws.rows) >= 2:
        try:
            headers = set([str(i.value).lower() for i in ws.rows[0]])
        except AttributeError:
            return False
        i = unicode(ws['A2'].value).strip().find('/')
        if VUID_HEADER_NAME_SET.issubset(headers) and i != -1:
            return True
    return False


def verify_vuid_headers_empty(vuid):
    wb = load_workbook(vuid.file.path)
    ws = wb.active
    try:
        headers = set([str(i.value).lower() for i in ws.rows[0]])
    except AttributeError:
        return False

    try:
        index = unicode(ws['A2'].value).strip().find('/')
        vuid_path = ws['A2'].value.strip()[index:]
    except AttributeError:
        return False

    return True


def verify_root_path(vuid):
    wb = load_workbook(vuid.file.path)
    ws = wb.active
    index = unicode(ws['A2'].value).strip().find('/')
    vuid_path = ws['A2'].value.strip()[index:]
    #print vuid_path, '-', vuid.project.root_path
    if vuid_path.startswith(vuid.project.root_path):
        return True
    else:
        return False


def verify_update_root_path(project, new_path):
    # if no vuids allow update root path
    if project.vuid_set.all().count() == 0:
        return True

    old_path = project.root_path
    if old_path.startswith(new_path):  # go up level, allowed
        try:
            with pysftp.Connection(project.bravo_server.address,
                                   username=project.bravo_server.account,
                                   private_key=settings.PRIVATE_KEY) as sftp:
                wc = sftp.execute('ls {0} -Rf | wc --l'.format(new_path))
                if int(wc[0]) > 15000:  # word count > 15k not allowed
                    return False
                else:
                    return True
        except (pysftp.ConnectionException,
                pysftp.CredentialException,
                pysftp.AuthenticationException,
                pysftp.SSHException):
            return False
    else:  # go deep level, not allowed
        return False
