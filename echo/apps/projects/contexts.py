from echo.apps.activity.models import Action
from echo.apps.projects.forms import ProjectForm, ServerForm, UploadForm, ProjectRootPathForm
from echo.apps.projects.models import Project, Language, VUID
from openpyxl import load_workbook
import unicodecsv


def context_home(user, sort=None):
    projects = Project.objects.filter(users__pk=user.pk)
    if sort:
        if sort == 'project_name':
            projects = projects.order_by('name')
            print type(projects)
        elif sort == '-project_name':
            projects = projects.order_by('-name')
            print type(projects)
        elif sort == 'total_prompts':
            projects = sorted(projects, key=lambda p: p.slots_total())
            print type(projects)
        elif sort == '-total_prompts':
            projects = sorted(projects, key=lambda p: p.slots_total(), reverse=True)
            print type(projects)
        elif sort == 'user_count':
            projects = sorted(projects, key=lambda p: p.users_total())
            print type(projects)
        elif sort == '-user_count':
            projects = sorted(projects, key=lambda p: p.users_total(), reverse=True)
            print type(projects)
    return {
        'feed': Action.objects.filter(
            scope__project__in=projects,
            type__in=[
                Action.ARCHIVE_PROJECT,
                Action.CREATE_PROJECT,
                Action.TESTER_JOIN_PROJECT,
                Action.TESTER_LEAVE_PROJECT,
                Action.UPLOAD_VUID]
        ).order_by('-time')[0:10],
        'projects': projects,
        'sort': sort
    }


def context_language(project, language_type='master'):
    return {
        'project': project,
        'language_type': language_type,
        'slots': project.voiceslots() if 'master' in language_type else project.voiceslots(
            filter_language=language_type)
    }


def context_language_csv(project, response, language_type='master'):
    response['Content-Disposition'] = 'attachment; filename="{0}_{1}.csv"'.format(project.name, language_type)
    writer = unicodecsv.writer(response)
    writer.writerow(["Name", "Path", "Language", "Verbiage", "History", "Status", "VUID Name", "VUID Date", "User"])
    if language_type == 'master':
        for slot in project.voiceslots().order_by('name'):
            writer.writerow(
                [slot.name, slot.path, slot.language.name, slot.verbiage, slot.history, slot.status, slot.vuid.filename,
                 slot.vuid.upload_date, slot.vuid.upload_by.username])
    else:
        for slot in project.voiceslots(filter_language=language_type).order_by('name'):
            writer.writerow(
                [slot.name, slot.path, slot.language.name, slot.verbiage, slot.history, slot.status, slot.vuid.filename,
                 slot.vuid.upload_date, slot.vuid.upload_by.username])
    return response


def context_new(project_form=ProjectForm()):
    return {'project_form': project_form}


def context_project(project, upload_form=UploadForm(), server_form=ServerForm(initial={'server': 0})):
    root_path_form = ProjectRootPathForm(initial={'root_path': project.root_path})
    return {
        'project': project,
        'languages': Language.objects.filter(project=project),
        'vuids': VUID.objects.filter(project=project),
        'upload_form': upload_form,
        'server_form': server_form,
        'root_path_form': root_path_form
    }


def context_projects(user, tab, sort=None, page=None):
    if tab == 'my':
        projects = Project.objects.filter(users__pk=user.pk, status=Project.TESTING)
    elif tab == 'archive':
        projects = Project.objects.filter(status=Project.CLOSED)
    else:
        tab = 'all'
        projects = Project.objects.all().exclude(users__pk=user.pk)

    if sort:
        if sort == 'project_name':
            projects = projects.order_by('name')
        elif sort == '-project_name':
            projects = projects.order_by('-name')
        elif sort == 'created_date':
            projects = sorted(projects, key=lambda p: p.created_date())
        elif sort == '-created_date':
            projects = sorted(projects, key=lambda p: p.created_date(), reverse=True)
        elif sort == 'last_modified':
            projects = sorted(projects, key=lambda p: p.last_modified_date())
        elif sort == '-last_modified':
            projects = sorted(projects, key=lambda p: p.last_modified_date(), reverse=True)
        elif sort == 'total_prompts':
            projects = sorted(projects, key=lambda p: p.slots_total())
        elif sort == '-total_prompts':
            projects = sorted(projects, key=lambda p: p.slots_total(), reverse=True)
        elif sort == 'user_count':
            projects = sorted(projects, key=lambda p: p.users_total())
        elif sort == '-user_count':
            projects = sorted(projects, key=lambda p: p.users_total(), reverse=True)

    return {
        'projects': projects,
        'tab': tab,
        'sort': sort
    }


def context_testslot(browser, project, slot, filepath):
    browser_type = None
    if browser:
        if browser.family and browser.version_string:
            browser_type = "{0}{1}".format(browser.family.lower(), browser.version_string.lower())
    return {
        'browser': browser_type,
        'project': project,
        'slot': slot,
        'file': filepath
    }


def context_vuid(vuid):
    wb = load_workbook(vuid.file.path)
    ws = wb.active
    return {
        'vuid': vuid,
        'headers': [i.value for i in ws.rows[0]],
        'path': ws['A2'].value,
        'records': [r for r in ws.rows[2:]]
    }




def context_temp(browser):
    browser_type = None
    if browser:
        if browser.family and browser.version_string:
            browser_type = "{0}{1}".format(browser.family.lower(), browser.version_string.lower())
    print browser_type
    return {
        'browser': browser_type
    }